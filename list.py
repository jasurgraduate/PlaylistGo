import sys
import os
import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QListWidget,
    QListWidgetItem,
    QVBoxLayout,
    QPushButton,
    QWidget,
    QLabel,
    QAbstractItemView,
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QCursor


class SongReorderWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Reorder Songs")
        self.setGeometry(100, 100, 800, 600)  # Larger window dimensions

        # Main layout
        self.main_layout = QVBoxLayout()

        # List widget for reordering songs
        self.reorder_list = QListWidget()
        self.reorder_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self.reorder_list.setStyleSheet(
            """
            QListWidget {
                background-color: #1e272e;
                color: #d2dae2;
                padding: 10px;
                border-radius: 10px;
            }
            QListWidget::item {
                padding: 10px;
                margin: 5px;
                border-radius: 5px;
                background-color: #485460;
            }
            QListWidget::item:hover {
                background-color: #006435;
            }
            QListWidget::item:selected {
                background-color: #ff60c0;
                color: #ffffff;
            }
        """
        )
        self.reorder_list.viewport().setCursor(QCursor(Qt.CursorShape.OpenHandCursor))
        self.reorder_list.viewport().installEventFilter(self)
        self.main_layout.addWidget(self.reorder_list)

        # Button to save the new order to Excel
        self.save_order_button = QPushButton("Save Order to Excel")
        self.save_order_button.setStyleSheet(
            """
            QPushButton {
                background-color: #3c40c6;
                color: #ffffff;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #575fcf;
            }
        """
        )
        self.save_order_button.clicked.connect(self.save_order_to_excel)
        self.main_layout.addWidget(self.save_order_button)

        # Label for displaying messages
        self.message_label = QLabel("")
        self.message_label.setStyleSheet(
            "font-size: 16px; color: #ffd32a; padding: 10px;"
        )
        self.main_layout.addWidget(self.message_label)

        # Set the main layout
        container = QWidget()
        container.setLayout(self.main_layout)
        self.setCentralWidget(container)

        # Load songs from Excel
        self.load_songs_from_excel()

    def load_songs_from_excel(self):
        excel_file = os.path.join(os.path.dirname(__file__), "songs.xlsx")
        try:
            df = pd.read_excel(excel_file, sheet_name="Active")
            for title in df["Title"]:
                list_item = QListWidgetItem(title)
                self.reorder_list.addItem(list_item)
        except FileNotFoundError:
            self.message_label.setText("Excel file not found! Please ensure it exists.")
        except Exception as e:
            self.message_label.setText(f"Failed to load songs: {e}")

    def save_order_to_excel(self):
        # Get the new order of songs from the reorder list
        new_order = [
            self.reorder_list.item(i).text() for i in range(self.reorder_list.count())
        ]

        # Use the existing Excel file
        excel_file = os.path.join(os.path.dirname(__file__), "songs.xlsx")
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook["Active"]
        except FileNotFoundError:
            self.message_label.setText("Excel file not found! Please ensure it exists.")
            return

        # Find the columns for "Title" and "YouTube Link"
        title_col = None
        link_col = None
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value == "Title":
                title_col = col
            elif cell_value == "YouTube Link":
                link_col = col

        if title_col is None or link_col is None:
            self.message_label.setText(
                "Could not find 'Title' or 'YouTube Link' columns in the Excel sheet."
            )
            return

        # Create a new DataFrame with the new order
        new_rows = []
        for title in new_order:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if row[title_col - 1].value == title:
                    new_rows.append([row[title_col - 1].value, row[link_col - 1].value])
                    break

        # Insert new rows at the top using pandas and openpyxl
        try:
            df = pd.read_excel(excel_file, sheet_name="Active")
            new_df = pd.DataFrame(new_rows, columns=["Title", "YouTube Link"])
            df = pd.concat([new_df, df], ignore_index=True)

            # Clear the existing sheet
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None

            # Write the updated DataFrame to the sheet
            for r_idx, row in enumerate(df.itertuples(index=False, name=None), 2):
                sheet.cell(row=r_idx, column=title_col, value=row[0])
                sheet.cell(row=r_idx, column=link_col, value=row[1])

            workbook.save(excel_file)
            self.message_label.setText("Order saved to Excel successfully!")
        except Exception as e:
            self.message_label.setText(f"Failed to save to Excel: {e}")

    def eventFilter(self, source, event):
        if (
            event.type() == event.Type.MouseButtonPress
            and source is self.reorder_list.viewport()
        ):
            source.setCursor(QCursor(Qt.CursorShape.ClosedHandCursor))
        elif (
            event.type() == event.Type.MouseButtonRelease
            and source is self.reorder_list.viewport()
        ):
            source.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))
        elif (
            event.type() == event.Type.MouseMove
            and source is self.reorder_list.viewport()
        ):
            if event.buttons() == Qt.MouseButton.LeftButton:
                source.setCursor(QCursor(Qt.CursorShape.ClosedHandCursor))
            else:
                source.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))
        return super().eventFilter(source, event)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = SongReorderWindow()
    window.show()
    sys.exit(app.exec())
